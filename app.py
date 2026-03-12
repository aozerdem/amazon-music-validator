import streamlit as st
import pandas as pd
import io
import re
import zipfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# --- CONFIG ---
THRESHOLDS = {"F": 100, "G": 41, "H": 41, "I": 41, "J": 41, "K": 41}
LANG_PATTERN = re.compile(r"^[a-zA-Z]{2,3}(-[a-zA-Z0-9]{2,4})?$")

st.set_page_config(page_title="Amazon Music ZIP Validator", layout="wide")

def is_lang_folder(name: str) -> bool:
    return bool(LANG_PATTERN.fullmatch(name))

def excel_len(v) -> int:
    return len(str(v)) if v is not None else 0

def process_excel(file_content, filename, lang):
    """Processes Excel content from the ZIP in memory."""
    # Use BytesIO to make the raw bytes readable by openpyxl
    file_stream = io.BytesIO(file_content)
    wb = load_workbook(file_stream, data_only=True)
    ws = wb.worksheets[0]
    
    violations = []
    target_letters = list(THRESHOLDS.keys())
    col_indexes = [column_index_from_string(l) for l in target_letters]

    for r in range(1, ws.max_row + 1):
        for c, col_letter in zip(col_indexes, target_letters):
            cell = ws.cell(row=r, column=c)
            actual = excel_len(cell.value)
            needed = THRESHOLDS[col_letter]
            if actual > needed:
                violations.append({
                    "language": lang,
                    "file": filename,
                    "row": r,
                    "column": col_letter,
                    "needed_length": needed,
                    "actual_length": actual,
                    "over_by": actual - needed,
                    "value": cell.value
                })
    return violations

# --- UI ---
st.title("📦 Amazon Music ZIP Validator")
st.markdown("Upload a **ZIP file** containing language folders (e.g., `de-DE/`, `fr-FR/`).")

uploaded_zip = st.file_uploader("Upload ZIP File", type=["zip"])

if uploaded_zip:
    all_violations = []
    files_processed = 0
    
    with zipfile.ZipFile(uploaded_zip) as z:
        # Get list of all files in the zip
        file_list = [f for f in z.namelist() if f.lower().endswith(('.xlsx', '.xlsm')) and not f.split('/')[-1].startswith('~$')]
        
        if not file_list:
            st.warning("No valid Excel files found in the ZIP.")
        else:
            progress_bar = st.progress(0)
            
            for idx, file_path in enumerate(file_list):
                # Logic: path/to/de-DE/file.xlsx -> folder is "de-DE"
                parts = file_path.split('/')
                
                # We look for a folder name in the path that matches our language pattern
                lang = "unknown"
                for p in parts:
                    if is_lang_folder(p):
                        lang = p
                        break
                
                filename = parts[-1]
                
                with z.open(file_path) as f:
                    content = f.read()
                    violations = process_excel(content, filename, lang)
                    all_violations.extend(violations)
                
                files_processed += 1
                progress_bar.progress((idx + 1) / len(file_list))

    if all_violations:
        df = pd.DataFrame(all_violations)
        st.error(f"⚠️ Found {len(df)} violations across {files_processed} files.")
        st.dataframe(df, use_container_width=True)

        # Generate custom filename logic
        first_orig_name = uploaded_zip.name[:10]
        clean_name = re.sub(r'[^a-zA-Z0-9]', '_', first_orig_name)
        export_filename = f"batch_lengthcheck_{clean_name}.xlsx"

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Violations")
        
        st.download_button(
            label="📥 Download Batch Report",
            data=output.getvalue(),
            file_name=export_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    elif files_processed > 0:
        st.balloons()
        st.success(f"✅ Scanned {files_processed} files. No violations found!")
